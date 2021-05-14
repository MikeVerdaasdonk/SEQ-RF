#!/usr/bin/env python
# coding: utf-8

# In[1]:


##package needed to do the exercise
import pandas as pd
from openpyxl import Workbook


# In[ ]:


df = pd.read_csv('C:/Users/s151675/Documents/Jaar_master_2/Kwartiel 3,4/Master project/data/201026 Aanlevering SEH/201026 Aanlevering SEH/201026 SEH Orders.csv',header=None,sep=None, engine='python')
# print(df)


# In[ ]:


rows = len(df)
column1 =[]
column2 =[]
column3 =[]
column4 =[]
column5 =[]
column6 =[]
for i in range(0,rows-1):
    column1.append(df.iloc[i,0])
    column2.append(df.iloc[i, 1])
    column3.append(str(df.iloc[i, 2]))
    column4.append(df.iloc[i, 3])
    column5.append(df.iloc[i, 4])
    date = df.iloc[i,3]
    print(date)
    time = df.iloc[i,4]
    date_string = str(date)
    time_string = str(time)

    new_string = date_string[0:11] + time_string[0:8]
    column6.append(new_string)
#



workbook = Workbook()
sheet = workbook.active
for j in range(0,rows-1):
    sheet.cell(j+1,1).value = column1[j]
    sheet.cell(j + 1, 2).value = column2[j]
    sheet.cell(j + 1, 3).value = column3[j]
    sheet.cell(j + 1, 4).value = column4[j]
    sheet.cell(j + 1, 5).value = column5[j]
    sheet.cell(j + 1, 6).value = column6[j]


workbook.save(filename='C:/Users/s151675/Documents/Jaar_master_2/Kwartiel 3,4/Master project/Data aangepast/test_new.xlsx')


# In[ ]:




